
# coding: utf-8

# In[15]:


from selenium import webdriver
from time import sleep
import requests
from lxml import etree
from PIL import Image
from hashlib import md5

driver = webdriver.Chrome()
driver.maximize_window()

driver.get('http://jxcp.zjiet.edu.cn/#/')
sleep(2)
#标签定位
search_input = driver.find_element_by_id('username')
#标签交互
search_input.send_keys('2018000013')
sleep(2)
#标签定位
search_input = driver.find_element_by_id('password')
#标签交互
search_input.send_keys('123456')
sleep(1)
#截图获取验证码图片
driver.save_screenshot('a.png')
imgelement = driver.find_element_by_xpath('//*[@id="root"]/div/div/div/div/form/div[3]/span[1]/img')   #定位标签
location = imgelement.location
size = imgelement.size
rangle = (int(location['x']), int(location['y']), int(location['x'] + size['width']),
          int(location['y'] + size['height']))  # 写成我们需要截取的位置坐标
i = Image.open("a.png")  # 打开截图
frame4 = i.crop(rangle)  # 使用Image的crop函数，从截图中再次截取我们需要的区域
frame4.save('save.png') # 保存我们接下来的验证码图片 进行打码
sleep(4)
#超级鹰验证码识别
class Chaojiying_Client(object):

    def __init__(self, username, password, soft_id):
        self.username = username
        password =  password.encode('utf8')
        self.password = md5(password).hexdigest()
        self.soft_id = soft_id
        self.base_params = {
            'user': self.username,
            'pass2': self.password,
            'softid': self.soft_id,
        }
        self.headers = {
            'Connection': 'Keep-Alive',
            'User-Agent': 'Mozilla/4.0 (compatible; MSIE 8.0; Windows NT 5.1; Trident/4.0)',
        }

    def PostPic(self, im, codetype):
        """
        im: 图片字节
        codetype: 题目类型 参考 http://www.chaojiying.com/price.html
        """
        params = {
            'codetype': codetype,
        }
        params.update(self.base_params)
        files = {'userfile': ('ccc.jpg', im)}
        result = requests.post('http://upload.chaojiying.net/Upload/Processing.php', data=params, files=files, headers=self.headers).json()
        return result

    def ReportError(self, im_id):
        """
        im_id:报错题目的图片ID
        """
        params = {
            'id': im_id,
        }
        params.update(self.base_params)
        r = requests.post('http://upload.chaojiying.net/Upload/ReportError.php', data=params, headers=self.headers)
        return r.json()

chaojiying = Chaojiying_Client('1551969060', 'luoli860826', '909607')	#用户中心>>软件ID 生成一个替换 96001
im = open('save.png', 'rb').read()													#本地图片文件路径 来替换 a.jpg 有时WIN系统须要//
result = chaojiying.PostPic(im, 1004)
print(result)
print('验证码结果为:' + result['pic_str'])
sleep(2)
#标签定位
search_input = driver.find_element_by_id('kaptcha')
#标签交互
search_input.send_keys(result['pic_str'])
#点击登录
btn = driver.find_element_by_css_selector('.ant-btn')
btn.click()
sleep(2)
#标签到同行评价
btn = driver.find_element_by_xpath('//*[@id="root"]/div/div[2]/div[1]/div[1]/div/div/ul/li/a')
btn.click()
#标签到同行评价
btn = driver.find_element_by_xpath('//*[@id="root"]/div/div[2]/div[1]/div[1]/div/div/ul/li/ul/li/a')
btn.click()
sleep(1)
driver.find_element_by_xpath("//div[@id='root']/div/div[3]/div[2]/div/form/div/div/div[4]/div[2]/div/span/div/div/span").click()
sleep(1)
driver.find_element_by_xpath("//div[@id='root']/div/div[3]/div[2]/div/form/div/div[2]/div[4]/div[2]/div/span/div/div/span").click()
sleep(1)
driver.find_element_by_xpath("//div[@id='root']/div/div[3]/div[2]/div/form/div/div[2]/div[3]/div[2]/div/span/div[2]/div/span").click()
sleep(1)
driver.find_element_by_xpath("//div[@id='root']/div/div[3]/div[2]/div/form/div/div[2]/div[3]/div[2]/div/span/div/div/span").click()
sleep(2)
#标签部门并填写
driver.find_element_by_xpath("//div[@id='root']/div/div[3]/div[2]/div/form/div/div/div/div[2]/div/span/span/span/span/span").click()
driver.find_element_by_xpath("//li/ul/li/span[2]/span").click()


# In[16]:


from lxml import etree 
import openpyxl as op
wb = op.load_workbook("E:\谷歌下载文件\自动化录音填表\自动听课笔记表.xlsx")
sh=wb["Sheet4"]
driver.find_element_by_xpath("//div[@id='root']/div/div[3]/div[2]/div/form/div/div/div[2]/div[2]/div/span/div/div/div/div").click()
sleep(2)
page_text = driver.page_source
selector = etree.HTML(page_text)
content = selector.xpath('/html/body/div[3]/div/div/div/ul/li')
Sheet_title_list = ["教师","班级","课程","周次","上课地点","开始节次","结束节次","星期"]
#教师列表
teacher_name_list = []
#班级列表
class_name_list = []
#课程列表
curriculum_name_list = []
#周次列表
week_name_list = []
#上课地点列表
class_location_list = []
#开始列表
start_section_list = []
week_list = ["星期一","星期二","星期三","星期四","星期五","星期六","星期日"]
for teacher_name in content:
    name = teacher_name.xpath("./text()")[0]
    teacher_name_list.append(name)
for start_section in range(1,16):
    start_section_list.append(start_section)
#循环遍历教师
for i in range(len(teacher_name_list)):
    if i != len(teacher_name_list)-1:
        driver.find_elements_by_class_name("ant-select-dropdown-menu-item")[i].click()
        sleep(2)
        #点击班级并获取班级信息
        driver.find_element_by_xpath("//*[@id='root']/div/div[3]/div[2]/div[1]/form/div/div[1]/div[3]/div[2]/div/span/div/div/div/div[1]").click()
        sleep(2)
        page_text = driver.page_source
        selector = etree.HTML(page_text)
        content = selector.xpath('/html/body/div[4]/div/div/div/ul/li')
        temporary_name_list = []
        for class_name in content:
            class_names = class_name.xpath("./text()")[0]
            if class_names != "无匹配结果":
                class_name_list.append(class_names)
                class_name_list = list(set(class_name_list))        
                temporary_name_list.append(class_names)
        #点击课程并获取课程信息
        driver.find_element_by_xpath("//*[@id='root']/div/div[3]/div[2]/div[1]/form/div/div[2]/div[1]/div[2]/div/span/div/div/div/div[1]").click()
        sleep(2)
        page_text = driver.page_source
        selector = etree.HTML(page_text)
        content = selector.xpath('/html/body/div[5]/div/div/div/ul/li')
        temporary_class_name_list = []
        for curriculum_name in content:
            curriculum_names = curriculum_name.xpath("./text()")[0]
            if curriculum_names != "无匹配结果":
                curriculum_name_list.append(curriculum_names)
                curriculum_name_list = list(set(curriculum_name_list))
                temporary_class_name_list.append(curriculum_names)
        #点击周次并获取第几周
        driver.find_element_by_xpath("//*[@id='root']/div/div[3]/div[2]/div[1]/form/div/div[1]/div[4]/div[2]/div/span/div/div/div/div").click()
        sleep(2)
        page_text = driver.page_source
        selector = etree.HTML(page_text)
        content = selector.xpath('/html/body/div[6]/div/div/div/ul/li')
        for week_name in content:
            week_names = week_name.xpath("./text()")[0]
            week_name_list.append(week_names)
            week_name_list = list(set(week_name_list))
        #上课地点点击
        driver.find_element_by_xpath("//*[@id='root']/div/div[3]/div[2]/div[1]/form/div/div[2]/div[2]/div[2]/div/span/div/div/div/div[1]").click()
        sleep(2)
        page_text = driver.page_source
        selector = etree.HTML(page_text)
        content = selector.xpath('/html/body/div[7]/div/div/div/ul/li')
        temporary_class_location_list = []
        for class_location in content:
            if class_location.xpath("./text()") != []:
                class_locations = class_location.xpath("./text()")[0]
                class_location_list.append(class_locations)
                class_location_list = list(set(class_location_list))
                temporary_class_location_list.append(class_locations)
        for a in range(1,9):
            sh.cell(1,a+i*9,Sheet_title_list[a-1])
        sh.cell(2,1+i*9,teacher_name_list[i])
        if temporary_name_list != []:
            for j in range(len(temporary_name_list)):
                sh.cell(j+2,2+i*9,temporary_name_list[j])
        else:
            sh.cell(2,2+i*9,"无匹配结果")
        if temporary_class_name_list != []:
            for j in range(len(temporary_class_name_list)):
                sh.cell(j+2,3+i*9,temporary_class_name_list[j])
        else:
            sh.cell(2,3+i*9,"无匹配结果")
        for j in range(len(week_name_list)):
            sh.cell(j+2,4+i*9,week_name_list[j])
        if temporary_class_location_list != []:
            for j in range(len(temporary_class_location_list)):
                sh.cell(j+2,5+i*9,temporary_class_location_list[j])
        else:
            sh.cell(2,5+i*9,"无匹配结果")
        for j in range(len(start_section_list)):
            sh.cell(j+2,6+i*9,start_section_list[j])
        for j in range(len(start_section_list)):
            sh.cell(j+2,7+i*9,start_section_list[j])
        for j in range(len(week_list)):
            sh.cell(j+2,8+i*9,week_list[j])
        wb.save("E:\谷歌下载文件\自动化录音填表\自动听课笔记表.xlsx")
        driver.find_element_by_xpath("//div[@id='root']/div/div[3]/div[2]/div/form/div/div/div[2]/div[2]/div/span/div/div/span").click()
        driver.find_element_by_xpath("//div[@id='root']/div/div[3]/div[2]/div/form/div/div/div[2]/div[2]/div/span/div/div/div/div").click()
    else:
        driver.find_elements_by_class_name("ant-select-dropdown-menu-item")[i].click()
        #点击班级并获取班级信息
        driver.find_element_by_xpath("//*[@id='root']/div/div[3]/div[2]/div[1]/form/div/div[1]/div[3]/div[2]/div/span/div/div/div/div[1]").click()
        sleep(2)
        page_text = driver.page_source
        selector = etree.HTML(page_text)
        content = selector.xpath('/html/body/div[4]/div/div/div/ul/li')
        temporary_name_list = []
        for class_name in content:
            class_names = class_name.xpath("./text()")[0]
            if class_names != "无匹配结果":
                class_name_list.append(class_names)
                class_name_list = list(set(class_name_list))
                temporary_name_list.append(class_names)
        #点击课程并获取课程信息
        driver.find_element_by_xpath("//*[@id='root']/div/div[3]/div[2]/div[1]/form/div/div[2]/div[1]/div[2]/div/span/div/div/div/div[1]").click()
        sleep(2)
        page_text = driver.page_source
        selector = etree.HTML(page_text)
        content = selector.xpath('/html/body/div[5]/div/div/div/ul/li')
        temporary_class_name_list = []
        for curriculum_name in content:
            curriculum_names = curriculum_name.xpath("./text()")[0]
            if curriculum_names != "无匹配结果":
                curriculum_name_list.append(curriculum_names)
                curriculum_name_list = list(set(curriculum_name_list))
                temporary_class_name_list.append(curriculum_names)
        #上课地点点击
        driver.find_element_by_xpath("//*[@id='root']/div/div[3]/div[2]/div[1]/form/div/div[2]/div[2]/div[2]/div/span/div/div/div/div[1]").click()
        sleep(2)
        page_text = driver.page_source
        selector = etree.HTML(page_text)
        content = selector.xpath('/html/body/div[7]/div/div/div/ul/li')
        temporary_class_location_list = []
        for class_location in content:
            if class_location.xpath("./text()") != []:
                class_locations = class_location.xpath("./text()")[0]
                class_location_list.append(class_locations)
                class_location_list = list(set(class_location_list))
                temporary_class_location_list.append(class_locations)
        for a in range(1,9):
            sh.cell(1,a+i*9,Sheet_title_list[a-1])
        sh.cell(2,1+i*9,teacher_name_list[i])
        if temporary_name_list != []:
            for j in range(len(temporary_name_list)):
                sh.cell(j+2,2+i*9,temporary_name_list[j])
        else:
            sh.cell(2,2+i*9,"无匹配结果")
        if temporary_class_name_list != []:
            for j in range(len(temporary_class_name_list)):
                sh.cell(j+2,3+i*9,temporary_class_name_list[j])
        else:
            sh.cell(2,3+i*9,"无匹配结果")
        for j in range(len(week_name_list)):
            sh.cell(j+2,4+i*9,week_name_list[j])
        if temporary_class_location_list != []:
            for j in range(len(temporary_class_location_list)):
                sh.cell(j+2,5+i*9,temporary_class_location_list[j])
        else:
            sh.cell(2,5+i*9,"无匹配结果")
        for j in range(len(start_section_list)):
            sh.cell(j+2,6+i*9,start_section_list[j])
        for j in range(len(start_section_list)):
            sh.cell(j+2,7+i*9,start_section_list[j])
        for j in range(len(week_list)):
            sh.cell(j+2,8+i*9,week_list[j])
        wb.save("E:\谷歌下载文件\自动化录音填表\自动听课笔记表.xlsx")
print(class_name_list)
print(curriculum_name_list)
print(week_name_list)
print(start_section_list)
print(week_list)


# In[44]:


from openpyxl  import Workbook
from openpyxl  import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation

wb = load_workbook(r'E:\谷歌下载文件\自动化录音填表\自动听课笔记表.xlsx')
print(wb.sheetnames)  
sheetnames = wb.get_sheet_names()
ws = wb.get_sheet_by_name(sheetnames[0])
print('打开文件')
# Create a data-validation object with list validation
teacher_A = ""
for i in teacher_name_list:
    if teacher_name_list.index(i) != len(teacher_name_list)-1:
        teacher_A = teacher_A + i + ","
    else:
        teacher_A = teacher_A + i
print(teacher_A)
dv = DataValidation(type="list", formula1='"%s"', allow_blank=True)%(teacher_A)
c1=ws['B1']
dv.add(c1)
ws.add_data_validation(dv)
wb.save(r'E:\谷歌下载文件\自动化录音填表\自动听课笔记表.xlsx')

